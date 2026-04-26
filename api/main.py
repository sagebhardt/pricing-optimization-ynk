"""
YNK Pricing Optimization API.

Serves the pricing dashboard: weekly actions, decisions, export, alerts,
audit log, feedback, admin panel. Authenticates via Google SSO.
"""

import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from fastapi import FastAPI, HTTPException, Query, Request, Body
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse, PlainTextResponse, JSONResponse
from pydantic import BaseModel
import pandas as pd
import json
from pathlib import Path
from datetime import datetime
from typing import Optional

app = FastAPI(
    title="YNK Pricing Optimization API",
    version="2.0",
    description="ML-driven pricing actions for Yaneken Retail Group",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.middleware("http")
async def auth_middleware(request: Request, call_next):
    """Protect API endpoints with Google OAuth2 token verification."""
    path = request.url.path

    # Public paths — no auth
    if path in ("/health", "/auth/config") or path.startswith("/assets/") or path in ("/favicon.svg", "/favicon.ico"):
        return await call_next(request)

    from config.auth import GOOGLE_CLIENT_ID, get_user_role

    # Dev mode — no auth configured
    if not GOOGLE_CLIENT_ID:
        request.state.user = {
            "email": "dev@local", "name": "Developer", "picture": "",
            "role": "admin", "permissions": ["approve", "audit", "export", "manage", "plan", "read"],
            "brands": None,
        }
        return await call_next(request)

    # Non-API routes — serve SPA (frontend handles login)
    api_prefixes = ("/pricing", "/decisions", "/export", "/alerts", "/model",
                    "/recommendations", "/sku/", "/audit", "/auth/me", "/admin", "/feedback",
                    "/analytics", "/estimate-impact", "/strategy", "/simulate", "/channel-stats")
    if not any(path.startswith(p) for p in api_prefixes):
        return await call_next(request)

    # Verify Google OAuth2 token
    auth_header = request.headers.get("Authorization", "")
    if not auth_header.startswith("Bearer "):
        return JSONResponse(status_code=401, content={"detail": "Authentication required"})

    try:
        from google.oauth2 import id_token
        from google.auth.transport import requests as google_requests
        idinfo = id_token.verify_oauth2_token(
            auth_header[7:], google_requests.Request(), GOOGLE_CLIENT_ID
        )
    except Exception:
        return JSONResponse(status_code=401, content={"detail": "Invalid or expired token"})

    email = idinfo.get("email", "")
    role_info = get_user_role(email)
    if role_info is None:
        return JSONResponse(status_code=403, content={"detail": f"Access denied for {email}"})

    request.state.user = {
        "email": email,
        "name": idinfo.get("name", email),
        "picture": idinfo.get("picture", ""),
        **role_info,
    }
    return await call_next(request)


BASE_DIR = Path(__file__).parent.parent




class DecisionPayload(BaseModel):
    brand: str
    week: str
    key: str
    status: Optional[str] = None
    manual_price: Optional[int] = None
    estimated_impact: Optional[dict] = None
    chain_scope: Optional[str] = None  # "all" | "ecomm" | "bm" — chain-wide decision
    # "store" (default, legacy) or "channel". Channel-grain decisions use keys
    # of the form "{parent_sku}-{bm|ecomm}" and go to decisions_channel_{week}.json;
    # chain_scope is mutually exclusive with grain=channel (the row IS the channel).
    grain: Optional[str] = "store"


class BulkDecisionPayload(BaseModel):
    brand: str
    week: str
    keys: list[str]
    status: str
    grain: Optional[str] = "store"


class ImpactEstimatePayload(BaseModel):
    brand: str
    parent_sku: str
    store: Optional[str] = None  # required when grain=store
    channel: Optional[str] = None  # required when grain=channel ("bm" or "ecomm")
    manual_price: int
    grain: Optional[str] = "store"


class SimulationPayload(BaseModel):
    brand: str
    discount_pct: float  # 0.0 - 0.50
    duration_weeks: int = 1
    filter_category: Optional[str] = None  # e.g. "Footwear", "Apparel"
    filter_vendor: Optional[str] = None  # e.g. "Nike", "Adidas"
    filter_skus: Optional[list[str]] = None  # specific parent SKUs
    filter_stores: Optional[list[str]] = None  # specific stores


class FeedbackPayload(BaseModel):
    brand: str
    week: str
    key: str
    implemented: bool
    actual_price: Optional[int] = None
    note: Optional[str] = ""


@app.get("/health")
def health():
    return {"status": "ok"}


@app.post("/ai/ask")
def ai_ask(request: Request, body: dict = Body(...)):
    """Ask the AI pricing assistant a question."""
    from api.ai_assistant import ask
    user = _get_user(request)
    question = body.get("question", "")
    brand = body.get("brand", "hoka")
    if not question:
        return {"error": "No question provided"}
    _check_brand_access(user, brand)
    answer = ask(question, brand)
    return {"question": question, "brand": brand, "answer": answer}


@app.get("/report/weekly")
def weekly_report(request: Request, brand: Optional[str] = Query(None)):
    """Generate weekly pricing report."""
    from api.weekly_report import generate_weekly_report, format_plain_text
    user = _get_user(request)
    report = generate_weekly_report(brand)
    return {"report": report, "text": format_plain_text(report)}


@app.get("/alerts")
def get_size_alerts(brand: Optional[str] = Query(None), min_attrition: float = Query(0.3)):
    """Get size curve depletion alerts. Optionally filter by brand."""
    from api import storage
    alerts = storage.load_alerts()
    if len(alerts) == 0:
        return {"alerts": [], "week": None, "total_alerts": 0}

    if brand and "brand" in alerts.columns:
        alerts = alerts[alerts["brand"] == brand.lower()]
        if len(alerts) == 0:
            return {"alerts": [], "week": None, "total_alerts": 0}

    latest = alerts[alerts["week"] == alerts["week"].max()]
    latest = latest[latest["attrition_rate"] >= min_attrition].sort_values("attrition_rate", ascending=False)

    results = []
    for _, row in latest.head(30).iterrows():
        results.append({
            "parent_sku": row["codigo_padre"],
            "store": row["centro"],
            "brand": row.get("brand", ""),
            "active_sizes": int(row["active_sizes_4w"]),
            "total_sizes": int(row["total_sizes_ever"]),
            "attrition_rate": round(row["attrition_rate"], 3),
            "core_completeness": round(row["core_completeness"], 3),
            "alert_reasons": row.get("alert_reasons", ""),
        })

    return {
        "week": str(alerts["week"].max().date()),
        "total_alerts": len(latest),
        "items": results,
    }


@app.get("/alerts/cross-store")
def get_cross_store_alerts(brand: Optional[str] = Query(None), min_price_spread: float = Query(0.0)):
    """Get cross-store pricing consistency alerts, enriched with product
    and store names looked up from the per-store pricing_actions CSV
    (which carries both `product` and `store_name` columns)."""
    from api import storage
    alerts = storage.load_cross_store_alerts(brand)
    if len(alerts) == 0:
        return {"week": None, "total_alerts": 0, "items": []}

    latest = alerts[alerts["week"] == alerts["week"].max()]
    if min_price_spread > 0:
        latest = latest[latest["price_spread"] >= min_price_spread]

    # Build name lookups from per-store actions (always store-grain — that's
    # the only CSV with both store_name and product fields).
    product_by_sku = {}
    store_name_by_code = {}
    if brand:
        ad = storage.load_pricing_actions(brand)
        for it in ad.get("items", []):
            sku = str(it.get("parent_sku", ""))
            store = str(it.get("store", ""))
            if sku and "product" in it and sku not in product_by_sku:
                product_by_sku[sku] = it.get("product", "")
            if store and "store_name" in it and store not in store_name_by_code:
                store_name_by_code[store] = it.get("store_name", store)

    # Group by parent SKU with nested stores
    items = []
    for parent, group in latest.groupby("codigo_padre"):
        first = group.iloc[0]
        stores = []
        for _, row in group.iterrows():
            store_code = str(row["centro"])
            store = {
                "store": store_code,
                "store_name": store_name_by_code.get(store_code, store_code),
                "channel": row.get("channel", "bm"),
                "price": int(row["avg_precio_final"]) if pd.notna(row.get("avg_precio_final")) else None,
                "discount_rate": round(row["discount_rate"], 3) if pd.notna(row.get("discount_rate")) else None,
            }
            if "stock_on_hand" in row.index and pd.notna(row.get("stock_on_hand")):
                store["stock_on_hand"] = int(row["stock_on_hand"])
            if "velocity_4w" in row.index and pd.notna(row.get("velocity_4w")):
                store["velocity_4w"] = round(row["velocity_4w"], 1)
            stores.append(store)

        items.append({
            "parent_sku": parent,
            "product": product_by_sku.get(str(parent), ""),
            "brand": first.get("brand", ""),
            "n_stores": int(first["n_stores"]),
            "price_spread": round(first["price_spread"], 3),
            "discount_spread": round(first["discount_spread"], 3),
            "sync_price": int(first["sync_price"]) if pd.notna(first.get("sync_price")) else None,
            "alert_reasons": first.get("alert_reasons", ""),
            "stores": stores,
        })

    items.sort(key=lambda x: -x["price_spread"])
    return {
        "week": str(latest["week"].max().date()) if len(latest) > 0 else None,
        "total_alerts": len(items),
        "items": items[:50],
    }


@app.get("/analytics/competitors/{brand}")
def get_competitor_analytics(brand: str, request: Request):
    """Advanced competitor pricing analytics."""
    from api import storage
    user = _get_user(request)
    _check_brand_access(user, brand)
    return storage.load_competitor_analytics(brand)


@app.get("/strategy/brief/{brand}")
def get_competitive_brief(brand: str, request: Request):
    """Full competitive intelligence brief: position map, movements, opportunities, threats."""
    from api import storage
    user = _get_user(request)
    _check_brand_access(user, brand)
    return storage.load_competitive_brief(brand)


@app.get("/strategy/opportunities/{brand}")
def get_competitive_opportunities(brand: str, request: Request, limit: int = Query(20)):
    """Actionable competitive opportunities ranked by estimated weekly margin impact."""
    from api import storage
    user = _get_user(request)
    _check_brand_access(user, brand)
    brief = storage.load_competitive_brief(brand)
    if not brief.get("available"):
        return {"brand": brand, "available": False, "opportunities": []}
    return {
        "brand": brand,
        "available": True,
        "total_opportunity_value": brief.get("total_opportunity_value", 0),
        "opportunities": brief.get("opportunities", [])[:limit],
    }


@app.get("/strategy/threats/{brand}")
def get_competitive_threats(brand: str, request: Request):
    """Competitive threats requiring response, ranked by severity."""
    from api import storage
    user = _get_user(request)
    _check_brand_access(user, brand)
    brief = storage.load_competitive_brief(brand)
    if not brief.get("available"):
        return {"brand": brand, "available": False, "threats": []}
    return {
        "brand": brand,
        "available": True,
        "critical_count": brief.get("critical_threat_count", 0),
        "total_risk_value": brief.get("total_risk_value", 0),
        "threats": brief.get("threats", []),
    }


@app.get("/strategy/movements/{brand}")
def get_price_movements(brand: str, request: Request):
    """Recent competitor price movements (drops, raises, promo starts/ends)."""
    from api import storage
    user = _get_user(request)
    _check_brand_access(user, brand)
    brief = storage.load_competitive_brief(brand)
    if not brief.get("available"):
        return {"brand": brand, "available": False, "movements": []}
    return {
        "brand": brand,
        "available": True,
        "period": brief.get("period", {}),
        "movement_summary": brief.get("movement_summary", {}),
        "movements": brief.get("movements", []),
    }


# ── Promotion Simulator ────────────────────────────────────────────────────

@app.post("/simulate/promotion")
def simulate_promotion_endpoint(payload: SimulationPayload, request: Request):
    """Simulate a promotional scenario and return projected impact.

    Filters pricing actions by category/vendor/SKUs/stores, then simulates
    the given discount across all matching items using elasticity + margin math.
    """
    from api import storage
    from api.simulator import simulate_promotion
    user = _get_user(request)
    _check_brand_access(user, payload.brand)

    if payload.discount_pct < 0 or payload.discount_pct > 0.50:
        raise HTTPException(400, "discount_pct must be between 0.0 and 0.50")

    # Load actions + elasticity
    actions_data = storage.load_pricing_actions(payload.brand)
    items = actions_data.get("items", [])
    if not items:
        raise HTTPException(404, f"No pricing actions for {payload.brand}")

    # Apply filters
    if payload.filter_category:
        cat = payload.filter_category.lower()
        items = [i for i in items if cat in str(i.get("primera_jerarquia", "")).lower()
                 or cat in str(i.get("segunda_jerarquia", "")).lower()]
    if payload.filter_vendor:
        vendor = payload.filter_vendor.lower()
        items = [i for i in items if vendor in str(i.get("vendor_brand", "")).lower()]
    if payload.filter_skus:
        sku_set = set(payload.filter_skus)
        items = [i for i in items if i.get("parent_sku") in sku_set]
    if payload.filter_stores:
        store_set = set(str(s) for s in payload.filter_stores)
        items = [i for i in items if str(i.get("store")) in store_set]

    if not items:
        raise HTTPException(404, "No items match the filter criteria")

    # Load elasticity (SKU-level from GCS/local)
    elasticity_map = _load_elasticity_map(payload.brand)

    # Load competitor prices for context
    comp_summary = storage.load_competitor_summary(payload.brand)
    comp_prices = {}
    for ci in comp_summary.get("items", []):
        prices = [c["price"] for c in ci.get("competitors", []) if c.get("price", 0) > 0]
        if prices:
            comp_prices[ci["parent_sku"]] = min(prices)

    result = simulate_promotion(items, elasticity_map, payload.discount_pct, payload.duration_weeks, comp_prices)
    result["brand"] = payload.brand
    result["filter"] = {
        "category": payload.filter_category,
        "vendor": payload.filter_vendor,
        "skus": len(payload.filter_skus) if payload.filter_skus else None,
        "stores": payload.filter_stores,
    }
    return result


@app.post("/simulate/optimal-discount")
def simulate_optimal_discount(payload: SimulationPayload, request: Request):
    """Find the discount level that maximizes weekly margin for the filtered items."""
    from api import storage
    from api.simulator import find_optimal_discount
    user = _get_user(request)
    _check_brand_access(user, payload.brand)

    actions_data = storage.load_pricing_actions(payload.brand)
    items = actions_data.get("items", [])
    if not items:
        raise HTTPException(404, f"No pricing actions for {payload.brand}")

    # Apply same filters
    if payload.filter_category:
        cat = payload.filter_category.lower()
        items = [i for i in items if cat in str(i.get("primera_jerarquia", "")).lower()
                 or cat in str(i.get("segunda_jerarquia", "")).lower()]
    if payload.filter_vendor:
        vendor = payload.filter_vendor.lower()
        items = [i for i in items if vendor in str(i.get("vendor_brand", "")).lower()]
    if payload.filter_skus:
        sku_set = set(payload.filter_skus)
        items = [i for i in items if i.get("parent_sku") in sku_set]
    if payload.filter_stores:
        store_set = set(str(s) for s in payload.filter_stores)
        items = [i for i in items if str(i.get("store")) in store_set]

    if not items:
        raise HTTPException(404, "No items match the filter criteria")

    elasticity_map = _load_elasticity_map(payload.brand)
    result = find_optimal_discount(items, elasticity_map, payload.duration_weeks)
    result["brand"] = payload.brand
    result["items_count"] = len(items)
    return result


def _load_elasticity_map(brand: str) -> dict:
    """Load per-SKU elasticity as a dict. Cached via storage layer."""
    from api import storage
    import io
    elasticity = {}

    if storage._use_gcs():
        try:
            import pandas as pd
            bucket = storage._get_bucket()
            blob = bucket.blob(f"models/{brand.lower()}/elasticity_by_sku.parquet")
            if blob.exists():
                df = pd.read_parquet(io.BytesIO(blob.download_as_bytes()))
                reliable = df[df["confidence"].isin(["high", "medium"])]
                elasticity = reliable.set_index("codigo_padre")["elasticity"].to_dict()
        except Exception:
            pass

    if not elasticity:
        from pathlib import Path
        import pandas as pd
        fp = Path(__file__).parent.parent / "data" / "processed" / brand.lower() / "elasticity_by_sku.parquet"
        try:
            df = pd.read_parquet(fp)
            reliable = df[df["confidence"].isin(["high", "medium"])]
            elasticity = reliable.set_index("codigo_padre")["elasticity"].to_dict()
        except FileNotFoundError:
            pass

    return elasticity


@app.get("/model/info")
def get_model_info(brand: Optional[str] = Query(None)):
    """Get model metadata and performance metrics per brand."""
    from api import storage
    meta = storage.load_model_info(brand or "hoka")

    cls = meta.get("classifier", {})
    reg = meta.get("regressor", {})

    return {
        "version": "parent",
        "brand": brand or "default",
        "classifier": {
            "avg_auc": cls.get("avg_auc"),
            "avg_precision": cls.get("avg_precision"),
            "n_features": cls.get("n_features"),
        },
        "regressor": {
            "avg_mae": reg.get("avg_mae"),
            "avg_r2": reg.get("avg_r2"),
        },
        "note": meta.get("note", ""),
    }


@app.get("/analytics/overview")
def get_analytics_overview(request: Request):
    """Cross-brand overview: summary metrics for all accessible brands."""
    from api import storage
    user = _get_user(request)
    user_brands = user.get("brands")

    all_brands = ["hoka", "bold", "bamers", "oakley", "belsport"]
    if user_brands:
        all_brands = [b for b in all_brands if b in user_brands]

    # Parallel brand loading (I/O-bound GCS reads)
    from concurrent.futures import ThreadPoolExecutor

    def _load_brand_data(bid):
        return bid, storage.load_pricing_actions(bid), storage.load_model_info(bid)

    with ThreadPoolExecutor(max_workers=5) as ex:
        brand_data = list(ex.map(_load_brand_data, all_brands))

    brands = []
    for bid, ad, meta in brand_data:
        items = ad.get("items", [])
        week = ad.get("week")

        dec_data = storage.load_decisions(bid, week) if week else {"decisions": {}}
        dec_map = dec_data.get("decisions", {})

        total = len(items)
        decided = 0
        approved = 0
        for item in items:
            key = f"{item.get('parent_sku')}-{item.get('store')}"
            dec = dec_map.get(key, {})
            status = dec.get("status", "") if isinstance(dec, dict) else (dec or "")
            if status:
                decided += 1
                if status in ("approved", "bm_approved", "manual", "bm_manual", "planner_approved"):
                    approved += 1

        total_rev_delta = 0
        total_margin_delta = 0
        thin_margin = 0
        increases = 0
        for item in items:
            try:
                total_rev_delta += int(item.get("rev_delta", 0) or 0)
                total_margin_delta += int(item.get("margin_delta", 0) or 0)
            except (ValueError, TypeError):
                pass
            try:
                mp = item.get("margin_pct")
                if mp and mp != "" and float(mp) < 20:
                    thin_margin += 1
            except (ValueError, TypeError):
                pass
            if item.get("action_type") == "increase":
                increases += 1

        cls = meta.get("classifier", {})
        reg = meta.get("regressor", {})
        holdout = reg.get("holdout") or {}

        brands.append({
            "brand": bid,
            "week": week,
            "total_actions": total,
            "decided": decided,
            "approved": approved,
            "pending": total - decided,
            "increases": increases,
            "decreases": total - increases,
            "rev_delta": total_rev_delta,
            "margin_delta": total_margin_delta,
            "thin_margin_count": thin_margin,
            "classifier_auc": cls.get("avg_auc"),
            "regressor_r2": reg.get("avg_r2"),
            "holdout_r2": holdout.get("r2"),
            "n_samples": reg.get("n_samples"),
        })

    return {"brands": brands}


@app.get("/analytics/{brand}")
def get_analytics(brand: str, request: Request):
    """Get analytics panel data for a brand: model health, elasticity, lifecycle, impact."""
    from api import storage
    user = _get_user(request)
    _check_brand_access(user, brand)

    # Load data from caches
    meta = storage.load_model_info(brand)
    actions_data = storage.load_pricing_actions(brand)
    items = actions_data.get("items", [])
    shap_cls = storage.load_shap_features(brand, "classifier")
    shap_reg = storage.load_shap_features(brand, "regressor")
    elasticity = storage.load_elasticity_summary(brand)

    # Model section
    cls = meta.get("classifier", {})
    reg = meta.get("regressor", {})
    cls_holdout = cls.get("holdout") or {}
    reg_holdout = reg.get("holdout") or {}

    modelo = {
        "classifier_auc": cls.get("avg_auc"),
        "classifier_ap": cls.get("avg_precision"),
        "regressor_r2": reg.get("avg_r2"),
        "regressor_mae_pp": round(reg.get("avg_mae", 0) * 100, 1),
        "n_samples": reg.get("n_samples"),
        "n_features": cls.get("n_features"),
        "holdout_auc": cls_holdout.get("auc"),
        "holdout_r2": reg_holdout.get("r2"),
        "holdout_mae_pp": round(reg_holdout.get("mae", 0) * 100, 1) if reg_holdout.get("mae") else None,
        "holdout_n_samples": reg_holdout.get("n_samples") or cls_holdout.get("n_samples"),
        "training_mode": meta.get("training_mode"),
        "classifier_shap": shap_cls[:5],
        "regressor_shap": shap_reg[:5],
    }

    # Lifecycle / urgency distribution (derived from actions)
    urgency_dist = {}
    action_type_dist = {"increase": 0, "decrease": 0}
    confidence_dist = {}
    for item in items:
        u = item.get("urgency", "LOW")
        urgency_dist[u] = urgency_dist.get(u, 0) + 1
        at = item.get("action_type", "decrease")
        action_type_dist[at] = action_type_dist.get(at, 0) + 1
        ct = item.get("confidence_tier", "LOW")
        confidence_dist[ct] = confidence_dist.get(ct, 0) + 1

    ciclo = {
        "total_actions": len(items),
        "urgency_dist": urgency_dist,
        "action_type_dist": action_type_dist,
        "confidence_dist": confidence_dist,
    }

    # Impact breakdown (derived from actions)
    store_impact = {}
    subcat_impact = {}
    vendor_impact = {}
    thin_margin_count = 0
    for item in items:
        rev = 0
        margin = 0
        try:
            rev = int(item.get("rev_delta", 0) or 0)
            margin = int(item.get("margin_delta", 0) or 0)
        except (ValueError, TypeError):
            pass
        try:
            mp = item.get("margin_pct")
            if mp and mp != "" and float(mp) < 20:
                thin_margin_count += 1
        except (ValueError, TypeError):
            pass

        sn = item.get("store_name") or item.get("store", "?")
        store_impact.setdefault(sn, {"store": item.get("store", ""), "store_name": sn, "rev_delta": 0, "margin_delta": 0, "count": 0})
        store_impact[sn]["rev_delta"] += rev
        store_impact[sn]["margin_delta"] += margin
        store_impact[sn]["count"] += 1

        sc = item.get("subcategory", "Other")
        subcat_impact.setdefault(sc, {"subcategory": sc, "rev_delta": 0, "margin_delta": 0, "count": 0})
        subcat_impact[sc]["rev_delta"] += rev
        subcat_impact[sc]["margin_delta"] += margin
        subcat_impact[sc]["count"] += 1

        vb = item.get("vendor_brand", "Other")
        vendor_impact.setdefault(vb, {"vendor_brand": vb, "rev_delta": 0, "margin_delta": 0, "count": 0})
        vendor_impact[vb]["rev_delta"] += rev
        vendor_impact[vb]["margin_delta"] += margin
        vendor_impact[vb]["count"] += 1

    impacto = {
        "by_store": sorted(store_impact.values(), key=lambda x: abs(x["rev_delta"]), reverse=True)[:10],
        "by_subcategory": sorted(subcat_impact.values(), key=lambda x: abs(x["rev_delta"]), reverse=True)[:10],
        "by_vendor_brand": sorted(vendor_impact.values(), key=lambda x: abs(x["rev_delta"]), reverse=True)[:10],
        "thin_margin_count": thin_margin_count,
    }

    # Competitor section
    comp_summary = storage.load_competitor_summary(brand)

    return {
        "brand": brand,
        "modelo": modelo,
        "elasticidad": elasticity,
        "ciclo_de_vida": ciclo,
        "impacto": impacto,
        "competencia": comp_summary,
        "prediccion_vs_real": _build_outcome_summary(brand),
    }


def _build_outcome_summary(brand: str) -> dict:
    """Build prediction-vs-actual summary from outcome results."""
    from api import storage
    df = storage.load_outcomes(brand)
    if df is None or len(df) == 0:
        return {"available": False}

    valid = df[df["data_quality"] == "normal"].copy()
    if len(valid) == 0:
        return {"available": False}

    decisions_evaluated = len(valid)
    vel_errors = valid["velocity_error_pct"].dropna()
    median_velocity_error_pct = round(float(vel_errors.median()), 1) if len(vel_errors) > 0 else None

    dir_vals = valid["direction_correct"].dropna()
    pct_direction_correct = round(float(dir_vals.mean()) * 100, 1) if len(dir_vals) > 0 else None

    # Average actual revenue lift vs baseline
    lifts = valid["actual_lift_vs_baseline"].dropna()
    avg_actual_rev_lift = round(float(lifts.mean()), 1) if len(lifts) > 0 else None

    # Lift capture rate: how much of predicted lift was actually realized
    both = valid.dropna(subset=["actual_lift_vs_baseline", "predicted_lift_vs_baseline"])
    both = both[both["predicted_lift_vs_baseline"].abs() > 0.1]  # avoid divide-by-near-zero
    if len(both) > 0:
        capture = (both["actual_lift_vs_baseline"] / both["predicted_lift_vs_baseline"]).clip(-5, 5)
        lift_capture_rate = round(float(capture.median()) * 100, 1)
    else:
        lift_capture_rate = None

    # Breakdown by confidence tier
    by_confidence = {}
    for tier, group in valid.groupby("confidence_tier"):
        tier_errs = group["velocity_error_pct"].dropna()
        tier_dirs = group["direction_correct"].dropna()
        by_confidence[str(tier)] = {
            "count": len(group),
            "median_velocity_error_pct": round(float(tier_errs.median()), 1) if len(tier_errs) > 0 else None,
            "pct_direction_correct": round(float(tier_dirs.mean()) * 100, 1) if len(tier_dirs) > 0 else None,
        }

    # Breakdown by action type
    by_action_type = {}
    for at, group in valid.groupby("action_type"):
        at_errs = group["velocity_error_pct"].dropna()
        at_dirs = group["direction_correct"].dropna()
        by_action_type[str(at)] = {
            "count": len(group),
            "median_velocity_error_pct": round(float(at_errs.median()), 1) if len(at_errs) > 0 else None,
            "pct_direction_correct": round(float(at_dirs.mean()) * 100, 1) if len(at_dirs) > 0 else None,
        }

    # Worst mis-predictions (for drill-down table)
    worst_candidates = valid.dropna(subset=["velocity_error_pct"]).copy()
    worst_candidates["_abs_err"] = worst_candidates["velocity_error_pct"].abs()
    worst = worst_candidates.nlargest(5, "_abs_err", keep="first")
    worst_items = []
    for _, row in worst.iterrows():
        worst_items.append({
            "parent_sku": row["parent_sku"],
            "store": row["store"],
            "predicted_velocity": row.get("predicted_velocity"),
            "actual_velocity": row.get("actual_velocity"),
            "velocity_error_pct": row.get("velocity_error_pct"),
            "confidence_tier": row.get("confidence_tier"),
            "decision_week": row.get("decision_week"),
        })

    return {
        "available": True,
        "decisions_evaluated": decisions_evaluated,
        "weeks_evaluated": int(valid["decision_week"].nunique()),
        "median_velocity_error_pct": median_velocity_error_pct,
        "pct_direction_correct": pct_direction_correct,
        "avg_actual_rev_lift": avg_actual_rev_lift,
        "lift_capture_rate": lift_capture_rate,
        "by_confidence": by_confidence,
        "by_action_type": by_action_type,
        "worst_predictions": worst_items,
    }


@app.get("/analytics/outcomes/{brand}")
def get_outcome_details(brand: str, request: Request):
    """Per-decision drill-down for prediction vs actual outcomes."""
    from api import storage
    user = _get_user(request)
    _check_brand_access(user, brand)

    df = storage.load_outcomes(brand)
    if df is None or len(df) == 0:
        return {"available": False, "items": []}

    # Return all rows, sorted by absolute velocity error (worst first)
    valid = df[df["data_quality"] == "normal"].copy()
    valid["_abs_err"] = valid["velocity_error_pct"].abs()
    valid = valid.sort_values("_abs_err", ascending=False).drop(columns=["_abs_err"])

    items = valid.head(200).to_dict(orient="records")
    # Clean NaN for JSON serialization
    for item in items:
        for k, v in item.items():
            if isinstance(v, float) and (v != v):  # NaN check
                item[k] = None

    return {
        "available": True,
        "total": len(valid),
        "items": items,
    }


def _validate_grain(grain: str) -> str:
    """Normalize and validate a grain value (raises 400 if invalid)."""
    g = (grain or "store").lower()
    if g not in ("store", "channel"):
        raise HTTPException(400, f"Invalid grain '{grain}' — must be 'store' or 'channel'")
    return g


def _load_actions_by_grain(brand: str, grain: str) -> dict:
    """Dispatch the right storage loader for the requested grain."""
    from api import storage
    if grain == "channel":
        return storage.load_pricing_actions_channel(brand)
    return storage.load_pricing_actions(brand)


def _load_decisions_by_grain(brand: str, week, grain: str) -> dict:
    from api import storage
    if grain == "channel":
        return storage.load_decisions_channel(brand, week)
    return storage.load_decisions(brand, week)


def _save_decisions_by_grain(data: dict, grain: str):
    from api import storage
    if grain == "channel":
        storage.save_decisions_channel(data)
    else:
        storage.save_decisions(data)


@app.get("/pricing-actions")
def get_pricing_actions(
    request: Request,
    brand: Optional[str] = Query(None),
    grain: str = Query("store", description="'store' (default, legacy) or 'channel'"),
):
    """Get the weekly pricing action list at the requested grain."""
    if not brand:
        return {"items": [], "week": None, "total": 0, "grain": grain}
    grain = _validate_grain(grain)
    user = _get_user(request)
    _check_brand_access(user, brand)
    return _load_actions_by_grain(brand, grain)


@app.get("/channel-stats/{brand}")
def get_channel_stats(brand: str, request: Request):
    """Gap stats (chain_uniform_profit vs sum_per_store_profit) from the
    latest channel_aggregation run. Used by the dashboard to judge whether
    the uniform-price constraint costs material profit."""
    from api import storage
    user = _get_user(request)
    _check_brand_access(user, brand)
    return storage.load_channel_aggregation_stats(brand)


# ── Authentication helpers ────────────────────────────────────────────────────

def _get_user(request: Request) -> dict:
    """Get authenticated user from request state (set by middleware)."""
    return getattr(request.state, "user", {
        "email": "unknown", "name": "Unknown", "role": "viewer",
        "permissions": [], "brands": None,
    })


def _check_brand_access(user: dict, brand: str):
    """Raise 403 if user doesn't have access to this brand."""
    user_brands = user.get("brands")
    if user_brands is not None and brand.lower() not in user_brands:
        raise HTTPException(403, f"No access to brand {brand}")


@app.get("/auth/config")
def auth_config():
    """Auth configuration (public — needed before login)."""
    from config.auth import GOOGLE_CLIENT_ID
    return {"client_id": GOOGLE_CLIENT_ID, "required": bool(GOOGLE_CLIENT_ID)}


@app.get("/auth/me")
def auth_me(request: Request):
    """Current user info."""
    return _get_user(request)


# ── Admin: user management ────────────────────────────────────────────────────

class UserPayload(BaseModel):
    email: str
    role: str
    brands: Optional[list[str]] = None
    name: Optional[str] = ""


@app.get("/admin/users")
def admin_list_users(request: Request):
    """List all configured users."""
    from api import storage
    user = _get_user(request)
    if "manage" not in user.get("permissions", []):
        raise HTTPException(403, "Admin access required")
    cfg = storage.load_user_config()
    return cfg


@app.post("/admin/users")
def admin_set_user(payload: UserPayload, request: Request):
    """Add or update a user's role."""
    from api import storage
    user = _get_user(request)
    if "manage" not in user.get("permissions", []):
        raise HTTPException(403, "Admin access required")
    if payload.role not in ("admin", "brand_manager", "planner", "viewer"):
        raise HTTPException(400, "Invalid role")

    cfg = storage.load_user_config()
    cfg.setdefault("users", {})[payload.email.lower().strip()] = {
        "role": payload.role,
        "brands": payload.brands if payload.role in ("brand_manager", "planner") else None,
        "name": payload.name or "",
    }
    storage.save_user_config(cfg)

    storage.append_audit({
        "brand": "_system",
        "user_email": user["email"],
        "user_name": user["name"],
        "action": "set_role",
        "key": payload.email.lower(),
        "detail": payload.role,
    })
    return {"ok": True, "total_users": len(cfg["users"])}


@app.delete("/admin/users")
def admin_delete_user(email: str = Query(...), request: Request = None):
    """Remove a user."""
    from api import storage
    user = _get_user(request)
    if "manage" not in user.get("permissions", []):
        raise HTTPException(403, "Admin access required")

    cfg = storage.load_user_config()
    removed = cfg.get("users", {}).pop(email.lower().strip(), None)
    if not removed:
        raise HTTPException(404, "User not found")
    storage.save_user_config(cfg)

    storage.append_audit({
        "brand": "_system",
        "user_email": user["email"],
        "user_name": user["name"],
        "action": "remove_user",
        "key": email.lower(),
    })
    return {"ok": True}


@app.put("/admin/domains")
def admin_set_domains(request: Request, domains: list[str]):
    """Update allowed email domains."""
    from api import storage
    user = _get_user(request)
    if "manage" not in user.get("permissions", []):
        raise HTTPException(403, "Admin access required")
    cfg = storage.load_user_config()
    cfg["allowed_domains"] = [d.lower().strip() for d in domains]
    storage.save_user_config(cfg)
    return {"ok": True}


# ── Manual Price Impact Estimation ────────────────────────────────────────────

@app.post("/estimate-impact")
def estimate_impact(payload: ImpactEstimatePayload, request: Request):
    """Recalculate expected velocity/revenue/margin for a manually-set price."""
    from api.pricing_math import estimate_manual_price_impact
    user = _get_user(request)
    _check_brand_access(user, payload.brand)

    grain = _validate_grain(payload.grain)
    actions_data = _load_actions_by_grain(payload.brand, grain)

    action = None
    if grain == "channel":
        if not payload.channel:
            raise HTTPException(400, "channel required when grain='channel' ('bm' or 'ecomm')")
        channel = payload.channel.lower()
        if channel not in ("bm", "ecomm"):
            raise HTTPException(400, "channel must be 'bm' or 'ecomm'")
        for item in actions_data.get("items", []):
            if item.get("parent_sku") == payload.parent_sku and str(item.get("channel")) == channel:
                action = item
                break
        if not action:
            raise HTTPException(404, f"Channel action not found: {payload.parent_sku} / {channel}")
    else:
        if not payload.store:
            raise HTTPException(400, "store required when grain='store'")
        for item in actions_data.get("items", []):
            if item.get("parent_sku") == payload.parent_sku and str(item.get("store")) == str(payload.store):
                action = item
                break
        if not action:
            raise HTTPException(404, f"Action not found: {payload.parent_sku} in store {payload.store}")

    # estimate_manual_price_impact reads current_price, current_list_price,
    # current_velocity, unit_cost from the action dict — these keys exist
    # in both per-store and per-channel rows, so no special-casing needed.
    elasticity = None
    result = estimate_manual_price_impact(action, payload.manual_price, elasticity)
    return result


# ── Decisions (storage-backed) ────────────────────────────────────────────────

@app.get("/decisions")
def get_decisions(
    brand: str = Query(...),
    week: Optional[str] = Query(None),
    grain: str = Query("store"),
):
    """Get decisions for a brand at the requested grain (latest week by default)."""
    grain = _validate_grain(grain)
    return _load_decisions_by_grain(brand, week, grain)


@app.post("/decisions")
def save_decision(payload: DecisionPayload, request: Request):
    """Save a single approve/reject decision."""
    from api import storage
    user = _get_user(request)
    if "approve" not in user.get("permissions", []):
        raise HTTPException(403, "Permission 'approve' required")
    _check_brand_access(user, payload.brand)

    grain = _validate_grain(payload.grain)
    if grain == "channel" and payload.chain_scope:
        # At channel grain the row IS the chain-channel action; chain_scope doesn't apply
        raise HTTPException(400, "chain_scope is not valid when grain='channel'")

    data = _load_decisions_by_grain(payload.brand, payload.week, grain)
    data["week"] = payload.week
    data["brand"] = payload.brand.lower()
    data.setdefault("decisions", {})

    if payload.status is None or payload.status == "":
        data["decisions"].pop(payload.key, None)
        action = "undo"
    else:
        record = {
            "status": payload.status,
            "timestamp": datetime.now().isoformat(),
            "user": user["email"],
        }
        if payload.manual_price is not None:
            record["manual_price"] = payload.manual_price
        if payload.estimated_impact is not None:
            record["estimated_impact"] = payload.estimated_impact
        if payload.chain_scope is not None and grain == "store":
            record["chain_scope"] = payload.chain_scope

        # Chain-scope fan-out (legacy per-store path only) — writes to individual
        # store keys that match the scope. At grain=channel this fan-out is not
        # needed; the key already represents the channel.
        if grain == "store" and payload.chain_scope and payload.status:
            from config.vendor_brands import is_ecomm_store
            if "-chain-" not in payload.key:
                raise HTTPException(400, f"Invalid chain key format: {payload.key}")
            ad = storage.load_pricing_actions(payload.brand)
            parent_sku = payload.key.rsplit("-chain-", 1)[0]
            changed = 0
            for item in ad.get("items", []):
                if str(item.get("parent_sku")) != parent_sku:
                    continue
                store = str(item.get("store"))
                is_ec = is_ecomm_store(store)
                if payload.chain_scope == "ecomm" and not is_ec:
                    continue
                if payload.chain_scope == "bm" and is_ec:
                    continue
                store_key = f"{parent_sku}-{store}"
                if store_key not in data["decisions"]:  # store-level takes priority
                    data["decisions"][store_key] = {**record, "chain_key": payload.key}
                    changed += 1
            data["decisions"][payload.key] = record
            action = f"chain_{payload.status}_{payload.chain_scope}"
        else:
            data["decisions"][payload.key] = record
            action = payload.status

    _save_decisions_by_grain(data, grain)
    storage.append_audit({
        "brand": payload.brand.lower(),
        "user_email": user["email"],
        "user_name": user["name"],
        "action": action,
        "key": payload.key,
        "week": payload.week,
        "grain": grain,
    })
    return {"ok": True, "total": len(data["decisions"]), "grain": grain}


@app.post("/decisions/bulk")
def bulk_decisions(payload: BulkDecisionPayload, request: Request):
    """Bulk approve/reject (only sets keys not already decided)."""
    from api import storage
    user = _get_user(request)
    if "approve" not in user.get("permissions", []):
        raise HTTPException(403, "Permission 'approve' required")
    _check_brand_access(user, payload.brand)

    grain = _validate_grain(payload.grain)
    data = _load_decisions_by_grain(payload.brand, payload.week, grain)
    data["week"] = payload.week
    data["brand"] = payload.brand.lower()
    data.setdefault("decisions", {})

    changed = 0
    for key in payload.keys:
        if key not in data["decisions"]:
            data["decisions"][key] = {
                "status": payload.status,
                "timestamp": datetime.now().isoformat(),
                "user": user["email"],
            }
            changed += 1

    _save_decisions_by_grain(data, grain)
    if changed:
        storage.append_audit({
            "brand": payload.brand.lower(),
            "user_email": user["email"],
            "user_name": user["name"],
            "action": f"bulk_{payload.status}",
            "count": changed,
            "week": payload.week,
            "grain": grain,
        })
    return {"ok": True, "total": len(data["decisions"]), "grain": grain}


# ── Planner Approval (two-step workflow) ──────────────────────────────────────

class PlannerDecisionPayload(BaseModel):
    brand: str
    week: str
    keys: list[str]
    status: str  # "planner_approved" or "planner_rejected"
    grain: Optional[str] = "store"


@app.post("/decisions/plan")
def planner_decide(payload: PlannerDecisionPayload, request: Request):
    """Planner approves or rejects BM-proposed decisions."""
    from api import storage
    user = _get_user(request)
    if "plan" not in user.get("permissions", []):
        raise HTTPException(403, "Permission 'plan' required")
    _check_brand_access(user, payload.brand)

    if payload.status not in ("planner_approved", "planner_rejected"):
        raise HTTPException(400, "Status must be 'planner_approved' or 'planner_rejected'")

    grain = _validate_grain(payload.grain)
    data = _load_decisions_by_grain(payload.brand, payload.week, grain)
    data["week"] = payload.week
    data["brand"] = payload.brand.lower()
    data.setdefault("decisions", {})

    bm_statuses = {"bm_approved", "bm_rejected", "bm_manual", "approved", "rejected", "manual"}
    changed = 0
    for key in payload.keys:
        rec = data["decisions"].get(key, {})
        current_status = rec.get("status", "")
        if current_status not in bm_statuses:
            continue  # skip items not yet BM-decided
        rec["bm_status"] = current_status  # preserve original BM decision
        rec["planner_status"] = payload.status
        rec["planner_user"] = user["email"]
        rec["planner_timestamp"] = datetime.now().isoformat()
        rec["status"] = payload.status
        data["decisions"][key] = rec
        changed += 1

    _save_decisions_by_grain(data, grain)
    if changed:
        storage.append_audit({
            "brand": payload.brand.lower(),
            "user_email": user["email"],
            "user_name": user["name"],
            "action": f"planner_{payload.status}",
            "count": changed,
            "week": payload.week,
            "grain": grain,
        })
    return {"ok": True, "changed": changed, "total": len(data["decisions"]), "grain": grain}


@app.get("/decisions/planner-queue")
def planner_queue(
    request: Request,
    brand: str = Query(...),
    grain: str = Query("store"),
):
    """Get items awaiting planner approval for a brand at the requested grain."""
    user = _get_user(request)
    if "plan" not in user.get("permissions", []):
        raise HTTPException(403, "Permission 'plan' required")
    _check_brand_access(user, brand)

    grain = _validate_grain(grain)
    ad = _load_actions_by_grain(brand, grain)
    items = ad.get("items", [])
    week = ad.get("week")

    dec_data = _load_decisions_by_grain(brand, week, grain)
    dec_map = dec_data.get("decisions", {})

    # Items that BMs have decided but planners haven't reviewed
    bm_decided_statuses = {"bm_approved", "bm_rejected", "bm_manual", "approved", "rejected", "manual"}
    planner_done_statuses = {"planner_approved", "planner_rejected"}

    queue = []
    for item in items:
        # Key format depends on grain:
        #   store   → {parent_sku}-{store}
        #   channel → {parent_sku}-{channel}  (channel in {'bm','ecomm'})
        # Defend against a missing channel/store field in the CSV — without this
        # the key would form as "{parent}-" and silently miss the decisions map.
        if grain == "channel":
            ch = item.get("channel") or ""
            if not ch:
                continue
            key = f"{item.get('parent_sku')}-{ch}"
        else:
            store = item.get("store") or ""
            if not store:
                continue
            key = f"{item.get('parent_sku')}-{store}"
        dec = dec_map.get(key, {})
        status = dec.get("status", "")
        if status in bm_decided_statuses and status not in planner_done_statuses:
            queue.append({
                **item,
                "decision_key": key,
                "bm_status": status,
                "bm_user": dec.get("user", ""),
                "bm_timestamp": dec.get("timestamp", ""),
                "manual_price": dec.get("manual_price"),
            })

    return {"brand": brand, "week": week, "total": len(queue), "items": queue, "grain": grain}


# ── Audit log ─────────────────────────────────────────────────────────────────

@app.get("/audit")
def get_audit(request: Request, brand: str = Query(...), limit: int = Query(100, ge=1, le=500)):
    """Get recent audit log entries for a brand."""
    from api import storage
    user = _get_user(request)
    if "audit" not in user.get("permissions", []):
        raise HTTPException(403, "Permission 'audit' required")
    return {"items": storage.load_audit(brand, limit)}


# ── Feedback (ops implementation tracking) ────────────────────────────────────

@app.get("/feedback")
def get_feedback(brand: str = Query(...), week: Optional[str] = Query(None)):
    """Get ops implementation feedback for a brand."""
    from api import storage
    return storage.load_feedback(brand, week)


@app.post("/feedback")
def save_feedback_item(payload: FeedbackPayload, request: Request):
    """Report whether a price change was implemented by ops."""
    from api import storage
    user = _get_user(request)

    data = storage.load_feedback(payload.brand, payload.week)
    data["week"] = payload.week
    data["brand"] = payload.brand.lower()
    data.setdefault("items", {})[payload.key] = {
        "implemented": payload.implemented,
        "actual_price": payload.actual_price,
        "note": payload.note or "",
        "reported_by": user["email"],
        "reported_at": datetime.now().isoformat(),
    }
    storage.save_feedback(data)

    storage.append_audit({
        "brand": payload.brand.lower(),
        "user_email": user["email"],
        "user_name": user["name"],
        "action": "feedback_implemented" if payload.implemented else "feedback_skipped",
        "key": payload.key,
        "week": payload.week,
    })
    return {"ok": True}


# ── Export ────────────────────────────────────────────────────────────────────

def _format_clp(n) -> str:
    """Format number as CLP: $36.990"""
    try:
        v = int(round(float(n)))
        formatted = f"{abs(v):,}".replace(",", ".")
        return f"-${formatted}" if v < 0 else f"${formatted}"
    except (ValueError, TypeError):
        return str(n)


@app.get("/export/price-changes")
def export_price_changes(
    request: Request,
    brand: str = Query(...),
    format: str = Query("excel", enum=["excel", "text"]),
    grain: str = Query("store"),
):
    """Export approved price changes as Excel or plain text."""
    from api import storage
    user = _get_user(request)
    if "export" not in user.get("permissions", []):
        raise HTTPException(403, "Permission 'export' required")
    _check_brand_access(user, brand)

    grain = _validate_grain(grain)
    ad = _load_actions_by_grain(brand, grain)
    if not ad.get("items"):
        raise HTTPException(404, "No pricing actions found")
    df = pd.DataFrame(ad["items"])
    week = ad["week"]

    dec_data = _load_decisions_by_grain(brand, week, grain)
    dec_map = dec_data.get("decisions", {})

    if grain == "channel":
        # Channel key: {parent_sku}-{bm|ecomm}
        # Drop rows with a missing channel value before key-building — they
        # can never match a decision and would otherwise produce a stray
        # "{parent}-nan" / "{parent}-" key that silently excludes the row.
        df = df[df["channel"].astype(str).isin(["bm", "ecomm"])].copy()
        df["_key"] = df["parent_sku"].astype(str) + "-" + df["channel"].astype(str)
    else:
        df = df[df["store"].astype(str).str.len() > 0].copy()
        df["_key"] = df["parent_sku"].astype(str) + "-" + df["store"].astype(str)
    # When planner approval is required, only planner_approved items export.
    # Legacy "approved"/"manual" are included for backward compat during rollout.
    require_planner = os.getenv("REQUIRE_PLANNER_APPROVAL", "").lower() in ("true", "1")
    if require_planner:
        exportable_statuses = {"planner_approved"}
    else:
        # Legacy "approved"/"manual" from old clients pass through; new bm_* statuses require planner
        exportable_statuses = {"approved", "manual", "planner_approved"}
    approved = df[df["_key"].apply(
        lambda k: dec_map.get(k, {}).get("status", "") in exportable_statuses
    )].copy()

    # For manual decisions, override the recommended_price with the manual price
    for idx, row in approved.iterrows():
        key = row["_key"]
        dec = dec_map.get(key, {})
        if dec.get("manual_price"):
            approved.at[idx, "recommended_price"] = dec["manual_price"]
            impact = dec.get("estimated_impact", {})
            if impact.get("margin_pct") is not None:
                approved.at[idx, "margin_pct"] = impact["margin_pct"]
            if impact.get("velocity") is not None:
                approved.at[idx, "expected_velocity"] = impact["velocity"]

    if len(approved) == 0:
        raise HTTPException(400, "No hay acciones aprobadas para exportar")

    increases = approved[approved["action_type"] == "increase"].copy()
    markdowns = approved[approved["action_type"] != "increase"].copy()

    storage.append_audit({
        "brand": brand.lower(),
        "user_email": user["email"],
        "user_name": user["name"],
        "action": f"export_{format}",
        "count": len(approved),
        "week": week,
        "grain": grain,
    })

    if format == "text":
        return _export_text(brand, week, increases, markdowns, grain=grain)
    return _export_excel(brand, week, increases, markdowns, grain=grain)


def _export_text(brand, week, increases, markdowns, grain="store"):
    """Plain text export for copy-paste into messaging."""
    lines = []
    total = len(increases) + len(markdowns)
    impact = int(
        (increases["rev_delta"].astype(float).sum() if len(increases) else 0)
        + (markdowns["rev_delta"].astype(float).sum() if len(markdowns) else 0)
    )

    lines.append(f"CAMBIOS DE PRECIO — {brand.upper()}")
    lines.append(f"Semana: {week}  |  Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    lines.append(f"Aprobados: {total} cambios  |  Impacto estimado: {_format_clp(impact)}/semana")
    lines.append("")

    if len(increases) > 0:
        lines.append(f"{'═' * 70}")
        lines.append(f"  SUBIR PRECIO ({len(increases)} productos)")
        lines.append(f"{'═' * 70}")
        lines.append("")
        lines.append(f"{'SKU':<16} {'PRODUCTO':<32} {'ANTES':>12} {'NUEVO':>12}")
        lines.append(f"{'─' * 16} {'─' * 32} {'─' * 12} {'─' * 12}")
        for _, r in increases.iterrows():
            lines.append(
                f"{str(r['parent_sku']):<16} "
                f"{str(r['product'])[:32]:<32} "
                f"{_format_clp(r['current_price']):>12} "
                f"{_format_clp(r['recommended_price']):>12}"
            )
        lines.append("")

    if len(markdowns) > 0:
        lines.append(f"{'═' * 70}")
        lines.append(f"  REBAJAS ({len(markdowns)} productos)")
        lines.append(f"{'═' * 70}")
        lines.append("")
        lines.append(f"{'SKU':<16} {'PRODUCTO':<28} {'ANTES':>12} {'NUEVO':>12} {'DCTO':>8}")
        lines.append(f"{'─' * 16} {'─' * 28} {'─' * 12} {'─' * 12} {'─' * 8}")
        for _, r in markdowns.iterrows():
            lines.append(
                f"{str(r['parent_sku']):<16} "
                f"{str(r['product'])[:28]:<28} "
                f"{_format_clp(r['current_price']):>12} "
                f"{_format_clp(r['recommended_price']):>12} "
                f"{str(r['recommended_discount']):>8}"
            )
        lines.append("")

    return PlainTextResponse("\n".join(lines))


def _export_excel(brand, week, increases, markdowns, grain="store"):
    """Excel export with formatted sheets."""
    import io
    from api import storage
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "openpyxl no instalado — pip install openpyxl")

    wb = Workbook()
    wb.remove(wb.active)

    hdr_font = Font(bold=True, size=11, color="FFFFFF")
    hdr_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    money_fmt = '#,##0'
    row_border = Border(bottom=Side(style='thin', color='E5E7EB'))

    def build_sheet(ws, title, rows, columns, col_widths):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
        c = ws.cell(row=1, column=1, value=title)
        c.font = Font(bold=True, size=14)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
        ws.cell(
            row=2, column=1,
            value=f"Semana: {week}  |  Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  {len(rows)} productos"
        ).font = Font(size=10, color="6B7280")
        for ci, (col_name, _) in enumerate(columns, 1):
            cell = ws.cell(row=4, column=ci, value=col_name)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal='center')
        money_keys = {'current_price', 'recommended_price', 'rev_delta', 'current_list_price',
                      'current_weekly_rev', 'expected_weekly_rev'}
        for ri, row_data in enumerate(rows, 5):
            for ci, (_, key) in enumerate(columns, 1):
                val = row_data.get(key, "")
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.border = row_border
                if key in money_keys:
                    try:
                        cell.value = int(float(val))
                        cell.number_format = money_fmt
                    except (ValueError, TypeError):
                        pass
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

    # Column definitions vary by grain: channel rows have 'channel' instead of 'store_name'.
    if grain == "channel":
        increase_cols = [
            ("SKU", "parent_sku"), ("Producto", "product"), ("Canal", "channel"),
            ("Tiendas", "n_stores"),
            ("Precio Actual", "current_price"), ("Precio Nuevo", "recommended_price"),
            ("Delta Rev/Sem", "rev_delta"),
        ]
        increase_widths = [18, 35, 10, 10, 15, 15, 15]
        markdown_cols = [
            ("SKU", "parent_sku"), ("Producto", "product"), ("Canal", "channel"),
            ("Tiendas", "n_stores"),
            ("Descuento", "recommended_discount"), ("Precio Actual", "current_price"),
            ("Precio Nuevo", "recommended_price"), ("Urgencia", "urgency"),
            ("Delta Rev/Sem", "rev_delta"),
        ]
        markdown_widths = [18, 30, 10, 10, 12, 15, 15, 12, 15]
    else:
        increase_cols = [
            ("SKU", "parent_sku"), ("Producto", "product"), ("Tienda", "store_name"),
            ("Precio Actual", "current_price"), ("Precio Nuevo", "recommended_price"),
            ("Delta Rev/Sem", "rev_delta"),
        ]
        increase_widths = [18, 35, 25, 15, 15, 15]
        markdown_cols = [
            ("SKU", "parent_sku"), ("Producto", "product"), ("Tienda", "store_name"),
            ("Descuento", "recommended_discount"), ("Precio Actual", "current_price"),
            ("Precio Nuevo", "recommended_price"), ("Urgencia", "urgency"),
            ("Delta Rev/Sem", "rev_delta"),
        ]
        markdown_widths = [18, 30, 25, 12, 15, 15, 12, 15]

    if len(increases) > 0:
        ws = wb.create_sheet("Subir Precio")
        build_sheet(ws, f"SUBIR PRECIO — {brand.upper()}", increases.to_dict('records'),
                    increase_cols, increase_widths)

    if len(markdowns) > 0:
        ws = wb.create_sheet("Rebajas")
        build_sheet(ws, f"REBAJAS — {brand.upper()}", markdowns.to_dict('records'),
                    markdown_cols, markdown_widths)

    buffer = io.BytesIO()
    wb.save(buffer)

    suffix = f"_{grain}" if grain == "channel" else ""
    filename = f"cambios_precio_{brand.lower()}{suffix}_{week}.xlsx"
    storage.save_export(brand, filename, buffer.getvalue())

    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# Serve React dashboard (must be after all API routes)
STATIC_DIR = Path(__file__).parent / "static"
if STATIC_DIR.exists():
    app.mount("/assets", StaticFiles(directory=STATIC_DIR / "assets"), name="assets")

    @app.get("/{full_path:path}")
    def serve_spa(full_path: str):
        """Serve React SPA for any non-API route."""
        file_path = STATIC_DIR / full_path
        if file_path.exists() and file_path.is_file():
            return FileResponse(file_path)
        return FileResponse(STATIC_DIR / "index.html")
